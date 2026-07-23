"""xlsx 청커 단위 테스트 — _cell / _to_markdown / chunk_xlsx.

is_table meta와 표 설명 병합은 retriever의 '표는 한 시트만' 필터·검색 보강이 의존한다.
"""
from pathlib import Path

import openpyxl
import pytest

from rag.xlsx_chunking import XLSX_MAX_ROWS, XlsxTooManyRows, _cell, _to_markdown, chunk_xlsx

CORPUS_XLSX = (Path(__file__).resolve().parent.parent
               / 'sample_docs' / 'corpus_v2' / 'homeplus' / 'homeplus_10_멤버십혜택표.xlsx')


class TestCell:
    def test_None은_빈칸_수치는_문자열화(self):
        assert _cell(None) == ''
        assert _cell(3000) == '3000'
        assert _cell(0.5) == '0.5'
        assert _cell('텍스트') == '텍스트'


class TestToMarkdown:
    def test_기본_표_형식(self):
        md = _to_markdown(['등급', '적립률'], [['VIP', '2.0']])
        assert md.splitlines() == [
            '| 등급 | 적립률 |',
            '| --- | --- |',
            '| VIP | 2.0 |',
        ]

    def test_짧은_행은_빈칸_패딩_긴_행은_잘림(self):
        md = _to_markdown(['a', 'b'], [['1'], ['1', '2', '3']])
        lines = md.splitlines()
        assert lines[2] == '| 1 |  |'
        assert lines[3] == '| 1 | 2 |'


class TestChunkXlsx:
    def test_시트당_청크1개_meta와_시트명(self):
        chunks = chunk_xlsx(CORPUS_XLSX)
        # 실제 시트명 고정 — meta==heading_path 자기참조 비교는 둘이 같이 틀려도 통과한다
        assert [c.heading_path for c in chunks] == [['등급기준'], ['등급별쿠폰'], ['포인트정책']]
        assert [c.meta for c in chunks] == [
            {'is_table': True, 'sheet': '등급기준'},
            {'is_table': True, 'sheet': '등급별쿠폰'},
            {'is_table': True, 'sheet': '포인트정책'},
        ]
        assert [c.chunk_index for c in chunks] == [0, 1, 2]

    def test_표_설명_병합(self):
        with_desc = chunk_xlsx(CORPUS_XLSX, description='멤버십 기준표')
        assert with_desc[0].text.startswith('[멤버십 기준표]\n')
        without = chunk_xlsx(CORPUS_XLSX)
        assert not without[0].text.startswith('[')              # 설명 없으면 표만

    def test_빈_시트_스킵(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '데이터'
        ws.append(['품목', '가격'])
        ws.append(['닭가슴살', 5900])
        wb.create_sheet('빈시트')                                # 값 없는 시트
        p = tmp_path / 't.xlsx'
        wb.save(p)
        chunks = chunk_xlsx(p)
        assert len(chunks) == 1 and chunks[0].heading_path == ['데이터']
        # 헤더 행(rows[0])이 표 첫 줄로 포함 — 첫 데이터 행이 헤더로 둔갑하는 회귀 방지 (뮤테이션 생존자)
        assert chunks[0].text.splitlines()[0] == '| 품목 | 가격 |'

    def test_행_상한_초과시_거절(self, tmp_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['col'])
        for i in range(XLSX_MAX_ROWS + 1):
            ws.append([i])
        p = tmp_path / 'big.xlsx'
        wb.save(p)
        with pytest.raises(XlsxTooManyRows):
            chunk_xlsx(p)

    def test_중간_빈_행은_스킵(self, tmp_path):
        # 빈 행이 표에 끼어도 데이터로 세지 않아야 함 (행 상한 오거절·빈 행 노이즈 방지)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['col'])
        ws.append([1])
        ws.append([None])
        ws.append([2])
        p = tmp_path / 'gap.xlsx'
        wb.save(p)
        chunks = chunk_xlsx(p)
        assert chunks[0].text.splitlines()[2:] == ['| 1 |', '| 2 |']

    def test_정확히_상한이면_허용(self, tmp_path):
        # off-by-one(> → >=)으로 정상 상한 파일이 거절되는 회귀 방지
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['col'])
        for i in range(XLSX_MAX_ROWS):
            ws.append([i])
        p = tmp_path / 'exact.xlsx'
        wb.save(p)
        assert len(chunk_xlsx(p)) == 1
